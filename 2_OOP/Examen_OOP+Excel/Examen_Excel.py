# Examen CRUD met Excel
# Basisbestand: consultants_data.xlsx

# Import
import openpyxl
# Alternatief: alleen de nodige functies:
# from openpyxl import load_workbook, Workbook
from colorama import Fore #nodig voor CRUD
from tabulate import tabulate

class ExcelFileActions:
    # Deze Excel-klasse leest altijd een bestaand bestand in. Louter wegschrijven kan dus niet.

    def __init__(self, file_path):
        """Initialiseer de klasse met het bestandspad van de Excel-file. Geeft terug of bestand bestaat"""
        self.file_path = file_path
        self.file_exists = False
        '''
        # File-check gebeurt nu in try/except met file_exists als output. Onderstaande code overbodig.
        # Check file first
        from os import path
        if not path.isfile(file_path):
            raise FileNotFoundError(f"Bestand '{file_path}' bestaat niet.")
        if not file_path.lower().endswith(('.xlsx', '.xlsm')):
            raise ValueError(
                f"Bestand '{file_path}' heeft een ongeldig formaat. Alleen .xlsx en .xlsm worden ondersteund.")
        '''
        # ? Hier al controleren dat excelbestand open staat?
        # Anderzijds lijkt dit geen probleem te geven op deze plaats.
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            self.file_exists = True
        except Exception as e:
            # print(e.args)
            if e.args[0] == 2:
                print(f"Fout: Het bestand '{file_path}' is niet gevonden.")
            elif 'not support' in e.args[0]:
                print(f"Fout: Het bestand '{file_path}' is geen .xlsx of .xlsm.")
            else:
                print(f"Fout: Het bestand '{file_path}' geeft foutmelding  {e}.")
            self.file_exists = False

    def __str__(self):
        return f"File exists is {self.file_exists}. Workbook tried to open with path: {self.file_path}."

    def check_file_status(self, methode):
        if methode == 1:
            while True:  # repeat until the try statement succeeds
                try:
                    myfile = open(self.file_path, "r+")  # or "a+", whatever you need
                    break  # exit the loop
                except IOError:
                    input("Could not open file! Please close Excel. Press Enter to retry.")
                    # restart the loop
        elif methode == 2:
            import os
            while True:  # repeat until the try statement succeeds
                try:
                    os.rename(self.file_path, self.file_path)
                    print("File is closed.")
                    break
                except OSError:
                    input("Could not open file! Please close Excel. Press Enter to retry.")
        else:
            print("Gekozen methode voor 'check_file_status' bestaat niet. Gebruik 1 of 2.")

    def toon_tabbladen(self):
        """Toon de namen van de tabbladen in het Excel-bestand."""
        return self.workbook.sheetnames

    def lees_data_actieve_tabblad(self):
        """Toon de gegevens van actieve tabblad."""
        worksheet = self.workbook.active
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(row)
        return data

    def toon_alle_kolomnamen(self, methode):
        worksheet = self.workbook.active
        if methode == 1:
            kolommen = [cell.value for cell in worksheet[1]]
        elif methode == 2:
            for row in worksheet.iter_rows(values_only=True):
                kolommen = list(row)
                break
        elif methode == 3:
            data = self.toon_data_actieve_tabblad()
            if data:
                for idx, col_name in enumerate(data[0], start=1):
                    print(f"Kolom {chr(64 + idx)}: {col_name}")
                    kolommen = col_name
            else:
                print("Geen data gelezen in bestand")
                kolommen = []
        else:
            print("Methode voor 'toon_alle_kolommen' is een andere waarde dan 1, 2 of 3: dat kan niet")
        return kolommen

    def toon_inhoud_specifieke_kolom(self, kolomnaam, methode):
        worksheet = self.workbook.active
        Kolommen = ConsultantBestand.toon_alle_kolomnamen(2)
        kolom_index = Kolommen.index(kolomnaam)
        # Creëer een set van de kolomdata
        if methode == 1:
            kolom_data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                kolom_data.append(row[kolom_index - 1])
        elif methode == 2:
            kolom_letter = chr(65+ kolom_index)
            print("Gezochte kolom is: ", kolom_letter)
            kolom_data = [cell.value for cell in worksheet[kolom_letter]]
            kolom_data = kolom_data[1:] # verwijder eerste element (ofwel de kolomnaam)
        else:
            print("Methode voor 'toon_inhoud_specifieke_kolom' is niet 1 of 2")
        # Filter dubbelen
        kolom_data = list(set(kolom_data))
        return kolom_data, kolom_index

    def opslaan_data_in_actief_tabblad(self, data):
        # Pas op: voegt data onder toe aan vorige data
        worksheet = self.workbook.active
        # Voeg de data toe aan het werkblad
        for row in data:
            worksheet.append(row)
        # Sla het Excel-bestand op
        self.workbook.save(self.file_path)
        print("Excel-bestand opgeslagen met de data in het actieve tabblad.")

    def vervang_data_in_actief_tabblad_behalve_header(self, Data):
        # Pas op: dit ledigt worksheet en schrijft dan de nieuwe data. De header blijft echter
        worksheet = self.workbook.active
        VorigeData = []
        for row in worksheet.iter_rows(values_only=True):
            VorigeData.append(row)
        Lengte_VorigeData = len(VorigeData)
        worksheet.delete_rows(2,Lengte_VorigeData-1)
        # Schrijf nu de nieuwe data
        for row in Data[1:]: # Sla header over
            worksheet.append(row)
        self.workbook.save(self.file_path)
        # Voeg de data toe aan het werkblad
        print("Excel-bestand opgeslagen met de data in het actieve tabblad.")

    def sorteer_op_data_in_tabblad_op_kolomnummer(self, tabblad, kolomnummer):
        # Functie 4: Sorteer op naam
        try:
            if tabblad in self.workbook.sheetnames:
                ws = self.workbook[tabblad]
                data = []
                for row in ws.iter_rows(min_row=2, values_only=True):  # Start bij rij 2 om headers over te slaan
                    data.append(row)

                data.sort(key=lambda x: x[kolomnummer])  # Sorteer op een kolom (kolomnummer, startend op 0)
                for i, row in enumerate(data, start=2):  # Schrijf gesorteerde data terug
                    for j, value in enumerate(row, start=1):
                        ws.cell(row=i, column=j, value=value)
                self.workbook.save(self.file_path)
                print("Tabblad is gesorteerd op naam (en opgeslagen).")
            else:
                print(f"Fout: Tabblad '{tabblad}' bestaat niet.")
        except Exception as e:
            print(f"Er is een fout opgetreden: {e}")

    def write_xlsfile_with_data(self, file_path_out, data):
        Nieuw_workbook = openpyxl.Workbook() #Let op hoofdletter!
        worksheet = Nieuw_workbook.active
        # Voeg de data toe aan het werkblad
        for row in data:
            worksheet.append(row)
        # Sla het Excel-bestand op
        try:
            Nieuw_workbook.save(file_path_out)
            Nieuw_workbook.close() # Lijkt logisch. Stond niet in voorbeeld.
            print(f"Excel-bestand {file_path_out} opgeslagen met de data in het actieve tabblad.")
        except Exception as e:
            print(e.args)
            if 'not support' in e.args[0]:
                print(f"Fout: Het bestand '{self.file_path}' is geen .xlsx of .xlsm.")
            else:
                print(f"Fout: Het bestand '{self.file_path}' geeft foutmelding  {e}.")
            self.file_exists = False

    def voeg_rij_toe_in_actief_tabblad(self):
        Data = self.lees_data_actieve_tabblad()
        worksheet = self.workbook.active
        NieuweRij = vraag_nieuwe_rij_op(Data) #Data moet mee om header en data-types af te leiden
        worksheet.append(NieuweRij)
        self.workbook.save(self.file_path)
        return NieuweRij

    def Verwijder_rij_in_actief_tabblad(self):
        Data = self.lees_data_actieve_tabblad()
        worksheet = self.workbook.active
        VerwijderdeRijIndex, VerwijderdeRij = vraag_te_verwijderen_rij_op(Data)  # Data moet mee om index af te leiden
        print("Verwijderde rij: ", VerwijderdeRij)
        if VerwijderdeRijIndex >= 0:
            worksheet.delete_rows(VerwijderdeRijIndex+1) # offset van Header
            self.workbook.save(self.file_path)
        return VerwijderdeRij

    def sorteer_data_in_actief_tabblad(self, Data, Index_SorteerKolom, order):
        worksheet = self.workbook.active
        Data_zonder_header = Data[1:]
        if order.lower()=='descendent':
            Descendent = True
        else: Descendent = False
        Data_zonder_header.sort(key=lambda x: x[Index_SorteerKolom], reverse=Descendent)  # Sorteer op de eerste kolom (Naam)
        print("Gesorteerde data: ", Data_zonder_header)
        # Hieronder wordt cel per cel overschreven. Andere methode is 'vervang_data...'
        for i, row in enumerate(Data_zonder_header, start=2):  # Schrijf gesorteerde data terug
            for j, value in enumerate(row, start=1):
                worksheet.cell(row=i, column=j, value=value)
        self.workbook.save(self.file_path)
        print(f"Tabblad is gesorteerd op kolom met index {Index_SorteerKolom}.")

    def sluit_Excel(self):
        self.workbook.close()

'''
Andere functies dan de Excel-klasse
'''
def Toon_data(Data):
    print(tabulate(Data[1:], headers=Data[0], tablefmt="fancy_grid", numalign="center"))

def Geef_kolomindex_voor_kolomnaam(Kolomnaam, Data):
    GezochteKolom = Kolomnaam
    Index_GezochteKolom = None
    # print(Data[0])
    for index, KolomNaam in enumerate(Data[0]):
        if KolomNaam == GezochteKolom:
            Index_GezochteKolom = index
            return Index_GezochteKolom, True
    if Index_GezochteKolom == None:
        print(f"Kolomnaam {GezochteKolom} is niet gevonden in de kop van de data. ")
        return Index_GezochteKolom, False  # Actie mislukt

def vraag_nieuwe_rij_op(Data):
    Header = Data[0]
    Types = [str(type(Element)) for Element in Data[1]]
    # print(Types)
    Nieuwe_rij = []
    Huidige_IDs = [element[0] for element in Data[1:]]
    print('Huidige IDs: ', Huidige_IDs)
    Hoogste_ID = max(Huidige_IDs)
    ID = max((len(Data)), Hoogste_ID+1) # len(data) meenemen is misschien overbodig
    print("Nieuwe ID wordt: ", ID)
    Nieuwe_rij.append(ID)
    for num, element in enumerate(Header[1:]):
        Invoer = input(f"Geef waarde voor {element}: ")
        if Types[num+1] == "<class 'int'>":  # '+1' want eerste element ID wordt overgeslagen
            Invoer = int(Invoer)
        if Types[num + 1] == "<class 'float'>":  # '+1' want eerste element ID wordt overgeslagen
            Invoer = float(Invoer)
        Nieuwe_rij.append(Invoer)
    print(f"U heeft opgegeven als activiteit: {Nieuwe_rij}")
    RijTuple = tuple(Nieuwe_rij)
    # print(RijTuple)
    return RijTuple

def vraag_te_verwijderen_rij_op(Data):
    Huidige_IDs = [element[0] for element in Data[1:]]
    try:
        Index_rij = int(input('Geef index van de te verwijderen rij: '))
    except:
        print("De invoer was geen geheel getal")
        return -1, None # Foute invoer
    if Index_rij in Huidige_IDs:
        for index, rij in enumerate(Data):
            if rij[0] == Index_rij:
                VerwijderdeRijIndex = index
                VerwijderdeRij = rij
        return VerwijderdeRijIndex, VerwijderdeRij
    else:
        print(f"De opgegeven index {Index_rij} bestaat niet.")
        return -1, None

def Pas_uurtarief_consultant_aan():
    ConsultantData = ConsultantBestand.lees_data_actieve_tabblad()
    # Toon_data(ConsultantData)
    # ConsultantData omzetten van list van tuples naar list van lists, want anders kunnen cellen niet worden aangepast
    ConsultantData = [list(Rij) for Rij in ConsultantData]
    # print("Check dat ConsultantData nu list van list is geworden: ", ConsultantData)
    ZoekKolomNaam = 'Id'
    KolomData, KolomIndex = ConsultantBestand.toon_inhoud_specifieke_kolom(ZoekKolomNaam , 2) # methode 1 blijkt fout want geeft rij
    print(f"(KolomData van {ZoekKolomNaam} is {KolomData}")
    # Vind index 'uurtarief'
    TeVervangenKolom = 'Uurtarief'
    Index_TeVervangenKolom = None
    # print(ConsultantData[0])
    for index, KolomNaam in enumerate(ConsultantData[0]): #Bestudeer header
        if KolomNaam == TeVervangenKolom:
            Index_TeVervangenKolom = index
            break
    if Index_TeVervangenKolom == None:
        print(f"Kolomnaam {TeVervangenKolom} is niet gevonden in de kop van de data. ")
        return False       # Actie mislukt
    # print(f"Om een idee te geven: volgende ID's bestaan: {KolomData}")
    GezochteID = int(input(f"Welke ID wilt u van uurtarief veranderen? : "))
    if GezochteID in KolomData:
        NieuwTarief = float(input("Wat is het nieuwe uurtarief? : "))
        for i, row in enumerate(ConsultantData[1:]):
            if row[KolomIndex] == GezochteID:
                print(f"Rij ID = {row[KolomIndex]} met Oorspronkelijk tarief: {ConsultantData[i+1][Index_TeVervangenKolom]} wordt vervangen door nieuw tarief {NieuwTarief}")
                ConsultantData[i+1][Index_TeVervangenKolom] = NieuwTarief # +1 want eerst header uitgesloten
        ConsultantBestand.vervang_data_in_actief_tabblad_behalve_header(ConsultantData)
        return True # Actie gelukt
    else:
        print(f"De gezochte {ZoekKolomNaam} {GezochteID} komt niet voor in de data. ")
        return False  # Actie mislukt

def Sorteer_uurprijs():
    ConsultantData = ConsultantBestand.lees_data_actieve_tabblad()
    GezochteKolom = 'Uurtarief'
    Index_GezochteKolom, Gevonden = Geef_kolomindex_voor_kolomnaam(GezochteKolom, ConsultantData)
    if Gevonden:
        ConsultantBestand.sorteer_data_in_actief_tabblad(ConsultantData, Index_GezochteKolom, 'descendent')
        return True  # Actie gelukt
    else: return False # Actie mislukt

def Schrijf_HRbestand():
    ConsultantData = ConsultantBestand.lees_data_actieve_tabblad()
    GezochteKolom = 'Domein'
    GezochteTerm = 'HR'
    KolomIndex, Gevonden = Geef_kolomindex_voor_kolomnaam(GezochteKolom, ConsultantData)
    if Gevonden:
        DataExtract = []
        DataExtract.append(list(ConsultantData[0])) # De header
        for row in ConsultantData[1:]:
            Term = row[KolomIndex]
            if Term.lower() == GezochteTerm.lower():
                DataExtract.append(row)
        print("DataExtract: ", DataExtract)
        NieuwBestand = r"HR-consultants.xlsx"
        ConsultantBestand.write_xlsfile_with_data (NieuwBestand, DataExtract)
        print(f"Bestand {NieuwBestand} is weggeschreven.")

def Vraag_keuze():
    print("Excel Manager:")
    print(Fore.GREEN+"Maak keuze uit volgende lijst:")
    print(Fore.RESET+"1.	Toon alle consultants")
    print("2.	Voeg een consultant toe")
    print("3.	Verwijder een consultant")
    print("4.	Pas het uurtarief van een consultant aan")
    print("5.	Sorteer de consultants op uurprijs van hoog naar laag")
    print("6.	Maak een apart Excelbestand met HR-consultants")
    print(Fore.RED+"7.   Stop")
    print(Fore.RESET + " ")
    try:
        Keuze_lokaal = int(input("Mijn keuze is: "))
    except:
        print("Foutieve keuze. Alleen getallen kunnen worden ingevoerd.")
        Keuze_lokaal = None
    print("Uw keuze is:  ", Keuze_lokaal)
    return Keuze_lokaal

"""
*** Hoofdgedeelte *** 
"""
# Initialisatie
Foute_keus = 0
Keuze = None

# Programma
if __name__ == "__main__":
    # Open bestand (de Excelklasse werkt alleen als het bestand bestaat)
    BestandRel = r"consultants_data.xlsx"
    ConsultantBestand = ExcelFileActions(BestandRel) # Open Excel-bestand als object
    print(ConsultantBestand)
    if ConsultantBestand.file_exists == True:
        # Gebruiker maakt een keuze
        while ((Keuze != 7) and (Foute_keus < 4)):
            Keuze = Vraag_keuze()
            if Keuze == 1:
                Foute_keus = 0
                ConsultantData = ConsultantBestand.lees_data_actieve_tabblad()
                print(ConsultantData)
                Toon_data(ConsultantData)
                # DataDict = Data_to_dict(ConsultantData)
                toets = input("Keuze 1 is afgelopen. Druk op Enter")
            elif Keuze == 2: # Voeg activiteit toe
                Foute_keus = 0
                # ConsultantData = ConsultantBestand.lees_data_actieve_tabblad()
                NieuweRij = ConsultantBestand.voeg_rij_toe_in_actief_tabblad()
                print("Consultantdata aangepast met : ", NieuweRij)
                toets = input("Keuze 2 is afgelopen. Druk op Enter")
            elif Keuze == 3: # Verwijder activiteit
                Foute_keus = 0
                VerwijderdeRij = ConsultantBestand.Verwijder_rij_in_actief_tabblad()
                print("Consultantdata aangepast door te verwijderen: ", VerwijderdeRij)
            elif Keuze == 4:  # Pas uurtarief van consultant aan
                Foute_keus = 0
                Gelukt = Pas_uurtarief_consultant_aan() # Geeft succes retour
                print("Het uurtarief is aangepast: ", Gelukt)
            elif Keuze == 5:  # Sorteer op uurprijs van hoog naar laag"
                Foute_keus = 0
                Sorteer_uurprijs()
            elif Keuze == 6:  # Schrijf consultants in HR naar nieuw bestand"
                Foute_keus = 0
                Schrijf_HRbestand()
            else:
                print("Geen opdracht of vraag tot beëindiging")
                Foute_keus += 1
        print("Programma beëindigd")

    # Hiermee eindigen
    if ConsultantBestand.file_exists == True:
        ConsultantBestand.sluit_Excel()

