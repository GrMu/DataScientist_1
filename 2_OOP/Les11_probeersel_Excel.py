# from openpyxl import Workbook
import openpyxl
def check_file_status_1(File):
    while True:   # repeat until the try statement succeeds
        try:
            myfile = open(File, "r+") # or "a+", whatever you need
            break                             # exit the loop
        except IOError:
            input("Could not open file! Please close Excel. Press Enter to retry.")
            # restart the loop

def check_file_status_2(File):
    import os
    while True:  # repeat until the try statement succeeds
        try:
            os.rename(File, File)
            print("File is closed.")
            break
        except OSError:
            input("Could not open file! Please close Excel. Press Enter to retry.")

Excelbestand = "data1.xlsx"

# check_file_status_1("data1.xlsx")
check_file_status_2(Excelbestand)

data = [[1, "A"], [2, "B"], [3, "C"], [4, "D"]]
# wb = Workbook()
wb = openpyxl.load_workbook(Excelbestand)
ws = wb.active
max_row = openpyxl
# max_row = sheet.max_row
rij = ["Bentje", ]
