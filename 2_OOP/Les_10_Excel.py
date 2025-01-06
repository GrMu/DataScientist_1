# from openpyxl import Workbook
import openpyxl
def check_file_status_1(File):
    while True:   # repeat until the try statement succeeds
        try:
            myfile = open(File, "r+") # or "a+", whatever you need
            break                             # exit the loop
        except IOError:
            input("Could not open file! Please close Excel. Press Enter to retry.")
            # restart the loop

def check_file_status_2(File):
    import os
    while True:  # repeat until the try statement succeeds
        try:
            os.rename(File, File)
            print("File is closed.")
            break
        except OSError:
            input("Could not open file! Please close Excel. Press Enter to retry.")

# check_file_status_1("data1.xlsx")
check_file_status_2("data1.xlsx")

data = [[1, "A"], [2, "B"], [3, "C"], [4, "D"]]
# wb = Workbook()
wb = openpyxl.load_workbook("data1.xlsx")
ws = wb.active

ws['A8'] = 42
'''
for row in data:
    print(row)
    ws.append(row)
wb.save("data1.xlsx") #"Data\data1.xlsx

ws['B10'] = 42
ws.append(data)
wb.save("data1.xlsx") #"Data\data1.xlsx
'''
for i, row in enumerate(data, start = 2):
    for j, cell in enumerate(row, start=2):
        ws.cell(i, j, cell)
wb.save("data1.xlsx") #"Data\data1.xlsx


