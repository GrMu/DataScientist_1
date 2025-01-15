# Proefexamen CRUD met Excel
# Basisbestand: teambuilding.xlsx

import openpyxl
# Alternatief: alleen de nodige functies:
# from openpyxl import load_workbook, Workbook

class ExcelFileActions:
    def __init__(self, file_path):
        """Initialiseer de klasse met het bestandspad van de Excel-file."""
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
'''
Aanmaken __repr__
Check bestand uit controller
'''

    def check_file_status_1(self):
        while True:  # repeat until the try statement succeeds
            try:
                myfile = open(self.file_path, "r+")  # or "a+", whatever you need
                break  # exit the loop
            except IOError:
                input("Could not open file! Please close Excel. Press Enter to retry.")
                # restart the loop

    def check_file_status_2(self):
        import os
        while True:  # repeat until the try statement succeeds
            try:
                os.rename(self.file_path, self.file_path)
                print("File is closed.")
                break
            except OSError:
                input("Could not open file! Please close Excel. Press Enter to retry.")

    def toon_tabbladen(self):
        """Toon de namen van de tabbladen in het Excel-bestand."""
        return self.workbook.sheetnames

    def toon_data_tabblad_na_opvragen(self):
        """Toon de gegevens van een gekozen tabblad."""
        tabblad_naam = input("Geef de naam van het tabblad: ")
        if tabblad_naam in self.workbook.sheetnames:
            sheet = self.workbook[tabblad_naam]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            return data
        else:
            return f"Tabblad '{tabblad_naam}' bestaat niet."

    def toon_data_actieve_tabblad(self):
        """Toon de gegevens van actieve tabblad."""
        worksheet = self.workbook.active
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(row)
        return data

    def toon_alle_kolomnamen(self, methode):
        worksheet = self.workbook.active
        if methode == 1:
            kolommen = [cell.value for cell in worksheet[1]]
        elif methode == 2:
            for row in worksheet.iter_rows(values_only=True):
                kolommen = list(row)
                break
        elif methode == 3:
            data = self.toon_data_actieve_tabblad()
            if data:
                for idx, col_name in enumerate(data[0], start=1):
                    print(f"Kolom {chr(64 + idx)}: {col_name}")
                    kolommen = col_name
            else:
                print("Geen data gelezen in bestand")
                kolommen = []
        else:
            print("Methode is een andere waarde dan 1, 2 of 3: dat kan niet")
        return kolommen

    def toon_inhoud_specifieke_kolom(self, kolomnaam, methode):
        worksheet = self.workbook.active
        Kolommen = TeamBestand.toon_alle_kolomnamen(2)
        kolom_index = Kolommen.index(kolomnaam) + 1
        # Creëer een set van de kolomdata
        if methode == 1:
            kolom_data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                kolom_data.append(row[kolom_index - 1])
        elif methode == 2:
            kolom_letter = chr(65+ kolom_index -1)
            print("Gezochte kolom is: ", kolom_letter)
            kolom_data = [cell.value for cell in worksheet[kolom_letter]]
            kolom_data = kolom_data[1:] # verwijder eerste element (ofwel de kolomnaam)
        else:
            print("Methode is niet 1 of 2")
        # Filter dubbelen
        kolom_data = list(set(kolom_data))
        return kolom_data

    def voeg_tabblad_toe(self, tabblad_naam, data):
        """Voeg een nieuw tabblad toe met gegeven naam en data (lijst van rijen)."""
        ''' 
        Met try functie zoals in f.5 in Excel_controller is het nog mooier
        '''
        if isinstance(data, list):
            sheet = self.workbook.create_sheet(title=tabblad_naam)
            for row in data:
                sheet.append(row)
            self.workbook.save(self.file_path)
            return f"Tabblad '{tabblad_naam}' is toegevoegd."
        else:
            return "De ingevoerde data is geen lijst van rijen."

    def opslaan_data_in_actief_tabblad(self, data):
        # Pas op: dit overschrijft huidige data in het bestand
        worksheet = self.workbook.active
        # Voeg de data toe aan het werkblad
        for row in data:
            worksheet.append(row)
        # Sla het Excel-bestand op
        bestand = self.file_path
        wb.save(bestand)
        print("Excel-bestand opgeslagen als ", bestand)

'''
!!! Voeg functies uit tabbladen_lezen toe !!

Voeg functie 4 sorteren in excel_controller toe, verwijder daarbij regel 1 (data[1:])


'''

        def sluit_Excel(self):
            self.workbook.close()

# Main

# Open bestand
BestandAbs = r"C:\Users\mulderg\PycharmProjects\DataScientist_1\2_OOP\Data\teambuilding_activiteiten.xlsx"
BestandRel = r"Data\teambuilding_activiteiten.xlsx"
TeamBestand = ExcelFileActions(BestandRel)
    # ("Data\teambuilding_activiteiten.xlsx")
# check_file_status_1("data1.xlsx")
TeamBestand.check_file_status_1()
# Lees data (en toon de tabbladen)
Tabs = TeamBestand.toon_tabbladen()
print("Aanwezige tabbladen: ", Tabs)
TeamData = TeamBestand.toon_data_actieve_tabblad()
print("Alle data in Teambestand: ", TeamData)
Kolommen = TeamBestand.toon_alle_kolomnamen(3) #geef methode (1,2,3) mee
print("Kolomnamen:", Kolommen)
Kolomnaam = "activiteit"
Kolom_data = TeamBestand.toon_inhoud_specifieke_kolom(Kolomnaam, 2)
print("Kolom ", Kolomnaam, "bevat volgende inhoud: ", Kolom_data)

'''
Zet worksheet-data om naar dictionary
'''


# Hiermee eindigen
TeamBestand.sluit_Excel()

