import openpyxl

def toon_alle_kolom_namen(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active
    kolommen = [cell.value for cell in sheet[1]]
    print("Kolomnamen:", kolommen)
    workbook.close()

def verwijder_kolom_met_naam(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active
    kolommen = [cell.value for cell in sheet[1]]
    toon_alle_kolom_namen(workbook_path)
    kolom_naam = input("Geef de naam van de kolom die verwijderd moet worden: ")

    if kolom_naam in kolommen:
        kolom_index = kolommen.index(kolom_naam) + 1
        max_row = sheet.max_row
        for row in range(1, max_row + 1):
            sheet.cell(row=row, column=kolom_index).value = None
        sheet.delete_cols(kolom_index)
        workbook.save(workbook_path)
        print(f"Kolom '{kolom_naam}' en de bijbehorende gegevens zijn verwijderd.")
    else:
        print(f"Kolom '{kolom_naam}' niet gevonden.")
    workbook.close()

def voeg_kolom_toe(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active
    kolom_naam = input("Geef de naam van de kolom die toegevoegd moet worden: ")

    sheet.insert_cols(sheet.max_column + 1)
    sheet.cell(row=1, column=sheet.max_column, value=kolom_naam)
    workbook.save(workbook_path)
    print(f"Kolom '{kolom_naam}' is toegevoegd.")
    workbook.close()

# Bestandspad
bestand_pad = 'personeel1.xlsx'

# Voorbeeld gebruik
while True:
    print("\nKies een optie:")
    print("1. Toon alle kolomnamen")
    print("2. Verwijder een kolom met naam")
    print("3. Voeg een kolom toe")
    print("4. Stop")

    keuze = input("Maak een keuze (1-4): ")

    if keuze == '1':
        toon_alle_kolom_namen(bestand_pad)
    elif keuze == '2':
        verwijder_kolom_met_naam(bestand_pad)
    elif keuze == '3':
        voeg_kolom_toe(bestand_pad)
    elif keuze == '4':
        print("Programma gestopt.")
        break
    else:
        print("Ongeldige keuze, probeer opnieuw.")
