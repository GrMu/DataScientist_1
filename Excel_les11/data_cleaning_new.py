import openpyxl

def toon_alle_kolom_namen(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active
    kolommen = [cell.value for cell in sheet[1]]
    print("Kolomnamen:", kolommen)
    workbook.close()

def verwijder_kolom_met_naam(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active
    kolommen = [cell.value for cell in sheet[1]]
    toon_alle_kolom_namen(workbook_path)
    kolom_naam = input("Geef de naam van de kolom die verwijderd moet worden: ")

    if kolom_naam in kolommen:
        kolom_index = kolommen.index(kolom_naam) + 1
        max_row = sheet.max_row
        for row in range(1, max_row + 1):
            sheet.cell(row=row, column=kolom_index).value = None
        sheet.delete_cols(kolom_index)
        workbook.save(workbook_path)
        print(f"Kolom '{kolom_naam}' en de bijbehorende gegevens zijn verwijderd.")
    else:
        print(f"Kolom '{kolom_naam}' niet gevonden.")
    workbook.close()

def voeg_kolom_toe(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active
    kolom_naam = input("Geef de naam van de kolom die toegevoegd moet worden: ")

    sheet.insert_cols(sheet.max_column + 1)
    sheet.cell(row=1, column=sheet.max_column, value=kolom_naam)
    workbook.save(workbook_path)
    print(f"Kolom '{kolom_naam}' is toegevoegd.")
    workbook.close()

def excel_van_afdeling(workbook_path):
    afdeling = input("Geef de afdeling waarvan je een Excel wilt maken: ")
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active

    nieuwe_workbook = openpyxl.Workbook()
    nieuwe_sheet = nieuwe_workbook.active

    headers = [cell.value for cell in sheet[1]]
    nieuwe_sheet.append(headers)

    afdeling_kolom = headers.index("afdeling") + 1

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[afdeling_kolom - 1] == afdeling:
            nieuwe_sheet.append(row)

    nieuwe_workbook.save(f"{afdeling}_personeel.xlsx")
    print(f"Excel bestand voor afdeling '{afdeling}' is opgeslagen als '{afdeling}_personeel.xlsx'.")
    workbook.close()

def excel_van_bedrijfswagen(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active

    nieuwe_workbook = openpyxl.Workbook()
    nieuwe_sheet = nieuwe_workbook.active

    headers = [cell.value for cell in sheet[1]]
    nieuwe_sheet.append(headers)

    bedrijfswagen_kolom = headers.index("bedrijfswagen") + 1

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[bedrijfswagen_kolom - 1] == "ja":
            nieuwe_sheet.append(row)

    nieuwe_workbook.save("bedrijfswagens_personeel.xlsx")
    print("Excel bestand met medewerkers met een bedrijfswagen is opgeslagen als 'bedrijfswagens_personeel.xlsx'.")
    workbook.close()

# Bestandspad
bestand_pad = 'personeel1.xlsx'

# Voorbeeld gebruik
while True:
    print("\nKies een optie:")
    print("1. Toon alle kolomnamen")
    print("2. Verwijder een kolom met naam")
    print("3. Voeg een kolom toe")
    print("4. Maak een Excel bestand van een afdeling")
    print("5. Maak een Excel bestand van medewerkers met een bedrijfswagen")
    print("6. Stop")

    keuze = input("Maak een keuze (1-6): ")

    if keuze == '1':
        toon_alle_kolom_namen(bestand_pad)
    elif keuze == '2':
        verwijder_kolom_met_naam(bestand_pad)
    elif keuze == '3':
        voeg_kolom_toe(bestand_pad)
    elif keuze == '4':
        excel_van_afdeling(bestand_pad)
    elif keuze == '5':
        excel_van_bedrijfswagen(bestand_pad)
    elif keuze == '6':
        print("Programma gestopt.")
        break
    else:
        print("Ongeldige keuze, probeer opnieuw.")
