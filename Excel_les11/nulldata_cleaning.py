import openpyxl

def load_workbook(file_path):
    """
    Load the workbook from the specified file path.
    """
    return openpyxl.load_workbook(file_path)

def toon_data(file_path):
    """
    Toon alle data in de gegeven Excel-bestand.
    """
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        print(row)


def toon_data_leeg(file_path):
    """
    Toon ID's van velden waar de cellen leeg zijn in de kolommen.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    leeg_ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header
        id_val = row[0]
        if any(cell is None or str(cell).strip() == '' for cell in row):
            leeg_ids.append(id_val)

    print("ID's met lege velden:", leeg_ids)


def niet_null(file_path, output_path):
    """
    Maak een nieuwe Excel met gegevens waar 'gemeente' niet leeg is.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    nieuwe_wb = openpyxl.Workbook()
    nieuwe_ws = nieuwe_wb.active

    nieuwe_ws.append([cell.value for cell in ws[1]])  # Kopieer header

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] and str(row[3]).strip():  # Controleer of 'gemeente' niet leeg is
            nieuwe_ws.append(row)

    nieuwe_wb.save(output_path)
    print(f"Nieuwe Excel is opgeslagen als: {output_path}")


# Bestandspaden aanpassen aan je situatie
file_path = 'nulldata.xlsx'
output_path = 'niet_null_data.xlsx'

# Gebruik de functies
print("Toon alle data:")
toon_data(file_path)

print("\nToon ID's met lege velden:")
toon_data_leeg(file_path)

print("\nMaak een nieuwe Excel met niet-lege gemeente:")
niet_null(file_path, output_path)
