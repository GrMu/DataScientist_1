from openpyxl import load_workbook, Workbook
import os


# Controleer of een Excel-bestand geldig is
def controleer_excel_bestand(file_name):
    if not os.path.isfile(file_name):
        raise FileNotFoundError(f"Bestand '{file_name}' bestaat niet.")
    if not file_name.lower().endswith(('.xlsx', '.xlsm')):
        raise ValueError(f"Bestand '{file_name}' heeft een ongeldig formaat. Alleen .xlsx en .xlsm worden ondersteund.")


# Functie 1: Lees een tabblad in en toon de data
def lees_bestand_in():
    bestand = input("Welk bestand wil je openen (data of data2)? ") + ".xlsx"
    try:
        controleer_excel_bestand(bestand)
        wb = load_workbook(bestand)
        print("Beschikbare tabbladen:")
        for sheet_name in wb.sheetnames:
            print(f"- {sheet_name}")

        tabblad = input("Voer de naam van het tabblad in: ")
        if tabblad in wb.sheetnames:
            ws = wb[tabblad]
            print("\nGegevens van het tabblad:")
            for row in ws.iter_rows(values_only=True):
                print(row)
        else:
            print(f"Fout: Tabblad '{tabblad}' bestaat niet.")
    except Exception as e:
        print(f"Er is een fout opgetreden: {e}")


# Functie 2: Voeg een rij toe
def voeg_rij_toe():
    bestand = input("Welk bestand wil je openen (data of data2)? ") + ".xlsx"
    try:
        controleer_excel_bestand(bestand)
        wb = load_workbook(bestand)
        tabblad = input("Welk tabblad wil je aanpassen? ")
        if tabblad in wb.sheetnames:
            ws = wb[tabblad]
            naam = input("Naam: ")
            leeftijd = int(input("Leeftijd: "))
            woonplaats = input("Woonplaats: ")
            ws.append([naam, leeftijd, woonplaats])
            wb.save(bestand)
            print("Rij is succesvol toegevoegd.")
        else:
            print(f"Fout: Tabblad '{tabblad}' bestaat niet.")
    except Exception as e:
        print(f"Er is een fout opgetreden: {e}")


# Functie 3: Verwijder een rij op basis van naam
def verwijder_rij():
    bestand = input("Welk bestand wil je openen (data of data2)? ") + ".xlsx"
    try:
        controleer_excel_bestand(bestand)
        wb = load_workbook(bestand)
        tabblad = input("Welk tabblad wil je aanpassen? ")
        if tabblad in wb.sheetnames:
            ws = wb[tabblad]
            naam = input("Naam van de persoon die je wilt verwijderen: ")
            for row in ws.iter_rows(min_row=2, values_only=False):  # Start bij rij 2 om headers over te slaan
                if row[0].value == naam:
                    ws.delete_rows(row[0].row)
                    wb.save(bestand)
                    print(f"Rij met naam '{naam}' is verwijderd.")
                    return
            print(f"Fout: Geen rij gevonden met naam '{naam}'.")
        else:
            print(f"Fout: Tabblad '{tabblad}' bestaat niet.")
    except Exception as e:
        print(f"Er is een fout opgetreden: {e}")


# Functie 4: Sorteer op naam
def sorteer_op_naam():
    bestand = input("Welk bestand wil je openen (data of data2)? ") + ".xlsx"
    try:
        controleer_excel_bestand(bestand)
        wb = load_workbook(bestand)
        tabblad = input("Welk tabblad wil je sorteren? ")
        if tabblad in wb.sheetnames:
            ws = wb[tabblad]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):  # Start bij rij 2 om headers over te slaan
                data.append(row)

            data.sort(key=lambda x: x[0])  # Sorteer op de eerste kolom (Naam)
            for i, row in enumerate(data, start=2):  # Schrijf gesorteerde data terug
                for j, value in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=value)

            wb.save(bestand)
            print("Tabblad is gesorteerd op naam.")
        else:
            print(f"Fout: Tabblad '{tabblad}' bestaat niet.")
    except Exception as e:
        print(f"Er is een fout opgetreden: {e}")


# Functie 5: Voeg een nieuw tabblad toe
def voeg_tabblad_toe():
    bestand = input("Welk bestand wil je openen (data of data2)? ") + ".xlsx"
    try:
        controleer_excel_bestand(bestand)
        wb = load_workbook(bestand)
        nieuw_tabblad = input("Naam van het nieuwe tabblad: ")
        if nieuw_tabblad not in wb.sheetnames:
            wb.create_sheet(title=nieuw_tabblad)
            wb.save(bestand)
            print(f"Tabblad '{nieuw_tabblad}' is toegevoegd.")
        else:
            print(f"Fout: Tabblad '{nieuw_tabblad}' bestaat al.")
    except Exception as e:
        print(f"Er is een fout opgetreden: {e}")

from openpyxl import load_workbook, Workbook

def voeg_tabladen_samen():
    bestand = input("Welk bestand wil je samenvoegen (data of data2)? ") + ".xlsx"
    try:
        controleer_excel_bestand(bestand)  # Controleer of het bestand geldig is
        wb = load_workbook(bestand)

        # Maak een nieuw Excel-bestand voor de samenvatting
        nieuw_bestand = "samenvatting.xlsx"
        nieuw_wb = Workbook()
        samenvatting_ws = nieuw_wb.active
        samenvatting_ws.title = "Samenvatting"

        # Voeg een header toe
        samenvatting_ws.append(["Tabblad", "Naam", "Leeftijd", "Woonplaats"])

        # Loop door alle tabbladen in het bestand
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):  # Start bij rij 2 om headers te negeren
                samenvatting_ws.append([sheet_name, *row])

        # Sla de samenvatting op
        nieuw_wb.save(nieuw_bestand)
        print(f"Tabbladen zijn samengevoegd in '{nieuw_bestand}'.")
    except Exception as e:
        print(f"Er is een fout opgetreden: {e}")


# Hoofdprogramma
# Hoofdprogramma
if __name__ == "__main__":
    print("Excel Manager")
    print("1) Lees een bestand/tabblad in")
    print("2) Voeg een rij toe")
    print("3) Verwijder een rij op basis van naam")
    print("4) Sorteer tabblad op naam")
    print("5) Voeg een nieuw tabblad toe")
    print("6) Voeg tabbladen samen")
    keuze = input("Maak een keuze (1-6): ")

    if keuze == "1":
        lees_bestand_in()
    elif keuze == "2":
        voeg_rij_toe()
    elif keuze == "3":
        verwijder_rij()
    elif keuze == "4":
        sorteer_op_naam()
    elif keuze == "5":
        voeg_tabblad_toe()
    elif keuze == "6":
        voeg_tabladen_samen()
    else:
        print("Ongeldige keuze!")
