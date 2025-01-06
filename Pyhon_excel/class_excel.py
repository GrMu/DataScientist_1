import openpyxl

class ExcelReader:
    def __init__(self, file_path):
        """Initialiseer de klasse met het bestandspad van de Excel-file."""
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)

    def toon_tabladen(self):
        """Toon de namen van de tabbladen in het Excel-bestand."""
        return self.workbook.sheetnames

    def toon_data_tablad(self):
        """Toon de gegevens van een gekozen tabblad."""
        tabblad_naam = input("Geef de naam van het tabblad: ")
        if tabblad_naam in self.workbook.sheetnames:
            sheet = self.workbook[tabblad_naam]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            return data
        else:
            return f"Tabblad '{tabblad_naam}' bestaat niet."

    def voeg_tablad_toe(self, tabblad_naam, data):
        """Voeg een nieuw tabblad toe met gegeven naam en data (lijst van rijen)."""
        if isinstance(data, list):
            sheet = self.workbook.create_sheet(title=tabblad_naam)
            for row in data:
                sheet.append(row)
            self.workbook.save(self.file_path)
            return f"Tabblad '{tabblad_naam}' is toegevoegd."
        else:
            return "De ingevoerde data is geen lijst van rijen."

e = ExcelReader("data2.xlsx")
print(e.toon_tabladen())