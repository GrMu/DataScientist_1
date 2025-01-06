from openpyxl import load_workbook

# Laad het Excel-bestand
wb = load_workbook("data.xlsx")
ws = wb.active

# Lees de data in een 2D-lijst
data = []
for row in ws.iter_rows(values_only=True):  # values_only=True retourneert alleen de celwaarden
    data.append(list(row))

# Toon de ingevulde kolomnamen
print("Ingevulde kolomnamen:")
if data:
    for idx, col_name in enumerate(data[0], start=1):
        print(f"Kolom {chr(64 + idx)}: {col_name}")

# Vraag de gebruiker om een kolom te kiezen
kolom_keuze = input("\nVoer de kolomletter in die je wilt afdrukken (bijv. 'A'): ").upper()

# Controleer of de ingevoerde kolom geldig is
kolom_index = ord(kolom_keuze) - 65  # Converteer kolomletter naar index
if 0 <= kolom_index < len(data[0]):
    # Print de waarden van de gekozen kolom
    print(f"\nWaarden in kolom {kolom_keuze}:")
    for row in data[1:]:  # Sla de header over
        print(row[kolom_index])
else:
    print(f"Fout: Kolom {kolom_keuze} bestaat niet.")
