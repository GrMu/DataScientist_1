import openpyxl


# Functie 1: Toon de waarde uit cel A1
def toon_data_uit_cel_a1():
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    waarde = ws['A1'].value
    print(f'Waarde in cel A1: {waarde}')


# Functie 2: Toon de waarde uit een specifieke cel
def toon_data_uit_cel(cel):
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    waarde = ws[cel].value
    print(f'Waarde in cel {cel}: {waarde}')


# Functie 3: Toon de gegevens uit een bereik van cellen
def toon_data_uitbereik(start, einde):
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    cellen = ws[start:einde]
    for rij in cellen:
        for cel in rij:
            print(cel.value, end=' ')
        print()


# Functie 4: Toon de gegevens van een specifieke rij
def toon_rij_nr(rij):
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    for cel in ws[rij]:
        print(cel.value, end=' ')
    print()


# Functie 5: Toon de gegevens van een specifieke kolom
def toon_kolom(kolom):
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    for cel in ws[kolom]:
        print(cel.value)


# Functie 6: Wijzig de waarde van een specifieke cel
def wijzig_cel(cel, nieuwe_waarde):
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    ws[cel] = nieuwe_waarde
    wb.save('data.xlsx')
    print(f'Waarde in cel {cel} is gewijzigd naar: {nieuwe_waarde}')


# Hoofdprogramma: Hier kun je de functies aanroepen voor gebruik
if __name__ == '__main__':
    # Voorbeeld van hoe je de functies kunt gebruiken
    print("Toon data uit cel A1:")
    toon_data_uit_cel_a1()

    print("\nToon data uit cel B2:")
    toon_data_uit_cel('B2')

    print("\nToon data uit bereik A1:B2:")
    toon_data_uitbereik('A1', 'B2')

    print("\nToon data uit rij 3:")
    toon_rij_nr(3)

    print("\nToon data uit kolom A:")
    toon_kolom('A')

    print("\nWijzig waarde in cel A1 naar 'Nieuwe waarde':")
    wijzig_cel('A1', 'Nieuwe waarde')

    # Na het uitvoeren van de bovenstaande functies, kun je het bestand controleren
    # om te zien dat de wijziging is opgeslagen.
