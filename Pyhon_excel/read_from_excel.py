from openpyxl import load_workbook

# Laad het Excel-bestand
wb = load_workbook("data.xlsx")
ws = wb.active

# Lees de data in een 2D-lijst
data = []
for row in ws.iter_rows(values_only=True):  # values_only=True retourneert alleen de celwaarden
    data.append(list(row))

# Toon de 2D-lijst
print(data)