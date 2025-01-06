from openpyxl import load_workbook


# Laad het Excel-bestand
def get_sheetnames(file_name):
    try:
        wb = load_workbook(file_name, read_only=True)  # Open in alleen-lezen modus
        return wb.sheetnames  # Retourneer een lijst van tabbladnamen
    except FileNotFoundError:
        print(f"Fout: Het bestand '{file_name}' is niet gevonden.")
        return []


# Voorbeeldgebruik
if __name__ == "__main__":
    excel_file = "data2.xlsx"  # Bestand om te lezen
    sheet_names = get_sheetnames(excel_file)

    if sheet_names:
        print("Tabbladen in het bestand:")
        for sheet in sheet_names:
            print(f"- {sheet}")
    else:
        print("Geen tabbladen gevonden of bestand bestaat niet.")
from openpyxl import load_workbook, Workbook

# Functie om een tabbladnaam te wijzigen
def verander_naam_tabblad(file_name, oude_naam, nieuwe_naam):
    try:
        wb = load_workbook(file_name)
        if oude_naam in wb.sheetnames:
            wb[oude_naam].title = nieuwe_naam
            wb.save(file_name)
            print(f"Tabblad '{oude_naam}' is hernoemd naar '{nieuwe_naam}'.")
        else:
            print(f"Fout: Tabblad '{oude_naam}' bestaat niet.")
    except FileNotFoundError:
        print(f"Fout: Bestand '{file_name}' is niet gevonden.")

# Functie om een tabblad te verwijderen
def verwijder_tabblad(file_name, tabblad_naam):
    try:
        wb = load_workbook(file_name)
        if tabblad_naam in wb.sheetnames:
            ws = wb[tabblad_naam]
            wb.remove(ws)
            wb.save(file_name)
            print(f"Tabblad '{tabblad_naam}' is verwijderd.")
        else:
            print(f"Fout: Tabblad '{tabblad_naam}' bestaat niet.")
    except FileNotFoundError:
        print(f"Fout: Bestand '{file_name}' is niet gevonden.")

# Functie om een nieuw tabblad toe te voegen
def voeg_tabblad_toe(file_name, tabblad_naam):
    try:
        wb = load_workbook(file_name)
        if tabblad_naam not in wb.sheetnames:
            wb.create_sheet(title=tabblad_naam)
            wb.save(file_name)
            print(f"Tabblad '{tabblad_naam}' is toegevoegd.")
        else:
            print(f"Fout: Tabblad '{tabblad_naam}' bestaat al.")
    except FileNotFoundError:
        # Als het bestand niet bestaat, maak een nieuw Excel-bestand
        print(f"Bestand '{file_name}' bestaat niet. Een nieuw bestand wordt aangemaakt.")
        wb = Workbook()
        wb.active.title = tabblad_naam
        wb.save(file_name)
        print(f"Nieuw bestand '{file_name}' is gemaakt met tabblad '{tabblad_naam}'.")

voeg_tabblad_toe("data2.xlsx","Test1")